"""Import drugs from Государственный реестр лекарственных средств (ГРЛС).

Официальный источник: https://grls.rosminzdrav.ru/GRLS.aspx
Скачайте выгрузку (Excel/CSV) и импортируйте:

  python manage.py import_grls_drugs --file data/imports/grls.xlsx
  python manage.py import_grls_drugs --file data/imports/grls.csv --apply
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from django.db import transaction

from apps.catalog.models import Drug

# Типичные заголовки выгрузки ГРЛС / открытых данных / ручного экспорта.
TRADE_KEYS = (
    "торговое наименование",
    "торговое_наименование",
    "торговое имя",
    "trade_name",
    "tradename",
    "название",
    "name",
)
MNN_KEYS = (
    "мнн",
    "млн",
    "мнн / группировочное (химическое) наименование",
    "млн / группировочное (химическое) наименование",
    "международное непатентованное наименование",
    "группировочное (химическое) наименование",
    "internationalname",
    "international_name",
    "inn",
)
FORM_KEYS = (
    "лекарственная форма",
    "форма выпуска",
    "formrelease",
    "form_release",
    "dosage_form",
    "form",
)
DOSE_KEYS = (
    "дозировка",
    "доза",
    "dosage",
    "содержание",
)
STATUS_KEYS = (
    "состояние",
    "статус",
    "status",
    "состояние регистрационного удостоверения",
)
MANUFACTURER_KEYS = (
    "производитель",
    "manufacturer",
    "stages",
)
HOLDER_KEYS = (
    "наименование держателя / владельца регистрационного удостоверения",
    "держатель ру",
    "владелец ру",
    "holder",
    "nameregcertificate",
)

ACTIVE_STATUS_TOKENS = (
    "действующ",
    "изменён",
    "изменен",
    "выдано по правилам еаэс",
    "на подтверждении",
    "в иностранных упаковках",
)

INACTIVE_STATUS_TOKENS = (
    "исключён",
    "исключен",
    "истёкш",
    "истекш",
    "приостановлен",
    "отменен",
    "отменён",
)


@dataclass
class GrlsImportStats:
    rows_total: int = 0
    unique_names: int = 0
    created: int = 0
    updated: int = 0
    skipped_inactive: int = 0
    skipped_empty: int = 0
    errors: list[str] = field(default_factory=list)


def _norm_header(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().casefold())


def _pick(row: dict[str, Any], keys: tuple[str, ...]) -> str:
    normalized = {_norm_header(k): ("" if v is None else str(v).strip()) for k, v in row.items()}
    for key in keys:
        value = normalized.get(_norm_header(key), "")
        if value:
            return value
    return ""


def _clean_trade_name(name: str) -> str:
    name = re.sub(r"\s+", " ", (name or "").strip())
    name = name.strip(" «»\"'")
    # Часто в ГРЛС: «ВАЛВИР» / «Валвир®» / «Валвир (Valvir)»
    name = name.replace("®", "").replace("™", "").strip()
    if len(name) > 255:
        name = name[:255].rstrip()
    return name


def _is_active_status(status: str) -> bool:
    status_l = (status or "").casefold()
    if not status_l:
        return True
    if any(token in status_l for token in INACTIVE_STATUS_TOKENS):
        return False
    if any(token in status_l for token in ACTIVE_STATUS_TOKENS):
        return True
    # Неизвестный статус — лучше оставить, чем потерять препарат
    return True


def _build_description(*, mnn: str, form: str, dose: str, manufacturer: str, holder: str) -> str:
    parts: list[str] = []
    if mnn:
        parts.append(f"МНН: {mnn}.")
    if form:
        parts.append(f"Лекарственная форма: {form}.")
    if dose:
        parts.append(f"Дозировка: {dose}.")
    if manufacturer:
        parts.append(f"Производитель: {manufacturer}.")
    if holder:
        parts.append(f"Держатель РУ: {holder}.")
    parts.append("Источник: Государственный реестр лекарственных средств (ГРЛС).")
    return " ".join(parts)[:2000]


def _build_dosage(*, form: str, dose: str) -> str:
    chunks = [x for x in (form, dose) if x]
    return ", ".join(chunks)[:255]


def iter_grls_rows(path: Path) -> Iterable[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        yield from _iter_csv(path)
        return
    if suffix in {".xlsx", ".xlsm"}:
        yield from _iter_xlsx(path)
        return
    raise ValueError(f"Неподдерживаемый формат файла: {suffix}. Нужен .csv или .xlsx")


# Минздрав open-data CSV (2015–2017): multiline-поля ломают DictReader.
# Якорь — страна производителя, сразу после неё идут tradename и МНН:
#   ...,Исландия,Валвир,,Валацикловир,"таблетки...
_OPEN_DATA_COUNTRIES = sorted(
    [
        "Россия",
        "Индия",
        "Германия",
        "Швейцария",
        "Словения",
        "Франция",
        "Китай",
        "Венгрия",
        "Великобритания",
        "Нидерланды",
        "Украина",
        "США",
        "Италия",
        "Израиль",
        "Республика Беларусь",
        "Беларусь",
        "Австрия",
        "Польша",
        "Чешская Республика",
        "Дания",
        "Бельгия",
        "Испания",
        "Сербия",
        "Турция",
        "Болгария",
        "Исландия",
        "Латвия",
        "Ирландия",
        "Финляндия",
        "Швеция",
        "Канада",
        "Япония",
        "Кипр",
        "Румыния",
        "Аргентина",
        "Вьетнам",
        "Норвегия",
        "Пакистан",
        "Люксембург",
        "Португалия",
        "Чехия",
        "Мексика",
        "Казахстан",
        "Грузия",
        "Египет",
        "Иран",
        "Куба",
        "Литва",
        "Эстония",
        "Греция",
        "Армения",
        "Австралия",
        "Бразилия",
        "Узбекистан",
        "Республика Корея",
        "Корея",
        "Хорватия",
        "Республика Хорватия",
        "Словацкая Республика",
        "Босния и Герцеговина",
        "Соединенное Королевство",
        "Республика Молдова",
        "Республика Македония",
        "Тайвань",
        "Саудовская Аравия",
        "Иордания",
        "Таиланд",
        "Перу",
        "КНР",
        "Пуэрто-Рико",
        "Индонезия",
        "Лихтенштейн",
        "Оман",
        "ЮАР",
        "Мальта",
        "Сингапур",
        "Новая Зеландия",
        "Азербайджан",
        "Южно-Африканская Республика",
        "Гонконг (КНР)",
        "Республика Куба",
        "Республика Армения",
        "Республика Казахстан",
        "Республика Польша",
        "Республика Болгария",
        "Республика Сербия",
        "Республика Эстония",
        "Эстонская Республика",
        "Югославия",
        "Македония",
        "Панама",
        "Соединенные Штаты",
        "Корея Южная",
    ],
    key=len,
    reverse=True,
)
_OPEN_DATA_COUNTRY_SET = {c.casefold() for c in _OPEN_DATA_COUNTRIES}
_OPEN_DATA_TRADE_RE = re.compile(
    r",(?P<country>"
    + "|".join(re.escape(c) for c in _OPEN_DATA_COUNTRIES)
    + r"),(?P<trade>[^,\r\n\"]+),(?P<field_a>[^,\r\n\"]*),(?P<field_b>[^,\r\n\"]*)",
    re.IGNORECASE,
)
_ADDRESS_JUNK_RE = re.compile(
    r"(?i)\b(street|avenue|road|str\.|rue |plaza|building|house|basel|budapest|"
    r"zagreb|diemen|ingelheim|freiburg)\b|^\d{2,}|\d{3,}\s"
)


def _iter_minzdrav_opendata(text: str) -> Iterable[dict[str, Any]]:
    """Разбор повреждённого CSV open-data ГРЛС: tradename сразу после страны."""
    for match in _OPEN_DATA_TRADE_RE.finditer(text):
        trade = (match.group("trade") or "").strip().strip('"')
        if len(trade) < 2 or trade.startswith('"'):
            continue
        if trade.casefold() in _OPEN_DATA_COUNTRY_SET:
            continue
        if _ADDRESS_JUNK_RE.search(trade):
            continue
        if not re.search(r"[A-Za-zА-Яа-яЁё]", trade):
            continue
        mnn = (match.group("field_a") or "").strip() or (match.group("field_b") or "").strip()
        yield {
            "country": match.group("country"),
            "tradename": trade,
            "internationalname": mnn,
            "formrelease": "",
            "состояние": "Действующий",
        }


def _iter_csv(path: Path) -> Iterable[dict[str, Any]]:
    raw = path.read_bytes()
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")

    # Detect delimiter
    sample = text[:4096]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","

    # Минздрав open-data: английские заголовки + часто пустой/сдвинутый tradename.
    header_line = text.splitlines()[0].casefold() if text else ""
    if "tradename" in header_line and "internationalname" in header_line:
        rows = list(_iter_minzdrav_opendata(text))
        nonempty = sum(1 for r in rows if (r.get("tradename") or "").strip())
        if nonempty > 0:
            yield from rows
            return

    # Обычный CSV/TSV (ручной экспорт ГРЛС, sample и т.п.)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    for row in reader:
        yield {str(k or ""): v for k, v in row.items()}


def _iter_xlsx(path: Path) -> Iterable[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "Для Excel нужен пакет openpyxl. Установите: pip install openpyxl\n"
            "Или сохраните ГРЛС как CSV и импортируйте --file ...csv"
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    keys = [str(cell or "").strip() for cell in header]
    for values in rows:
        yield {keys[i]: values[i] for i in range(len(keys)) if keys[i]}


def aggregate_grls_drugs(path: Path) -> tuple[dict[str, dict[str, str]], GrlsImportStats]:
    """Собрать уникальные торговые названия → поля Drug."""
    stats = GrlsImportStats()
    by_name: dict[str, dict[str, str]] = {}

    for row in iter_grls_rows(path):
        stats.rows_total += 1
        trade = _clean_trade_name(_pick(row, TRADE_KEYS))
        if not trade:
            stats.skipped_empty += 1
            continue
        status = _pick(row, STATUS_KEYS)
        if not _is_active_status(status):
            stats.skipped_inactive += 1
            continue

        mnn = _pick(row, MNN_KEYS)
        form = _pick(row, FORM_KEYS)
        dose = _pick(row, DOSE_KEYS)
        manufacturer = _pick(row, MANUFACTURER_KEYS)
        holder = _pick(row, HOLDER_KEYS)
        description = _build_description(
            mnn=mnn, form=form, dose=dose, manufacturer=manufacturer, holder=holder
        )
        dosage = _build_dosage(form=form, dose=dose)

        key = trade.casefold()
        existing = by_name.get(key)
        if existing is None:
            by_name[key] = {
                "name": trade,
                "description": description,
                "dosage": dosage,
                "mnn": mnn,
            }
        else:
            # Более полное описание / МНН побеждает
            if len(description) > len(existing.get("description") or ""):
                existing["description"] = description
            if dosage and (not existing.get("dosage") or len(dosage) > len(existing["dosage"])):
                existing["dosage"] = dosage
            if mnn and not existing.get("mnn"):
                existing["mnn"] = mnn

    stats.unique_names = len(by_name)
    return by_name, stats


@transaction.atomic
def import_grls_drugs(path: Path, *, dry_run: bool = True, limit: int = 0) -> GrlsImportStats:
    by_name, stats = aggregate_grls_drugs(path)
    items = list(by_name.values())
    if limit > 0:
        items = items[:limit]

    for item in items:
        name = item["name"]
        description = item.get("description") or ""
        dosage = item.get("dosage") or ""
        existing = Drug.objects.filter(name__iexact=name).first()
        if existing:
            changed = False
            # Не затираем богатые Vidal/curated описания короткими ГРЛС-строками
            if description and (
                not existing.description
                or existing.description.startswith("Источник:")
                or (
                    "ГРЛС" in existing.description
                    and len(description) > len(existing.description)
                )
                or (len(existing.description) < 40 and len(description) > len(existing.description))
            ):
                if not dry_run:
                    existing.description = description
                changed = True
            if dosage and not existing.dosage:
                if not dry_run:
                    existing.dosage = dosage
                changed = True
            if not existing.is_active:
                if not dry_run:
                    existing.is_active = True
                changed = True
            if changed:
                if not dry_run:
                    existing.save()
                stats.updated += 1
            continue

        if dry_run:
            stats.created += 1
            continue
        Drug.objects.create(
            name=name,
            description=description,
            dosage=dosage,
            instructions="",
            is_active=True,
        )
        stats.created += 1

    if dry_run:
        transaction.set_rollback(True)
    return stats
